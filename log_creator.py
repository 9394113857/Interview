import os
import logging
import sqlite3
from openpyxl import load_workbook, Workbook

# Function to create a directory if it doesn't exist
def create_directory(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)

# Create separate directories for log, excel, and database files
log_directory = 'logs'
excel_directory = 'excel_files'
db_directory = 'database_files'

create_directory(log_directory)
create_directory(excel_directory)
create_directory(db_directory)

# Function to create the log file
def create_log_file(file_path):
    if not os.path.exists(file_path):
        with open(file_path, 'w'):
            pass

# Function to perform mathematical operations
def perform_operation(a, b, operation):
    if operation == '+':
        result = a + b
        logging.info(f"Addition: {a} + {b} = {result}")
    elif operation == '-':
        result = a - b
        logging.info(f"Subtraction: {a} - {b} = {result}")
    elif operation == '*':
        result = a * b
        logging.info(f"Multiplication: {a} * {b} = {result}")
    elif operation == '/':
        if b != 0:
            result = a / b
            logging.info(f"Division: {a} / {b} = {result}")
        else:
            logging.error("Cannot divide by zero")
            result = None
    else:
        logging.error("Invalid operation")
        result = None

    return result

# Function to create SQLite table
def create_table(conn):
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS Operations
                      (id INTEGER PRIMARY KEY AUTOINCREMENT,
                       operation TEXT,
                       result REAL)''')
    conn.commit()

# Function to insert operation and result into SQLite table
def insert_operation(conn, operation, result):
    cursor = conn.cursor()
    cursor.execute('''INSERT INTO Operations (operation, result) VALUES (?, ?)''', (operation, result))
    conn.commit()
    logging.info(f"Operation inserted into SQLite table: {operation} - Result: {result}")

# Path to the log file
log_file_path = os.path.join(log_directory, 'mylog.log')

# Create the log file if it doesn't exist
create_log_file(log_file_path)

# Configure logging to append to the existing log file
logging.basicConfig(filename=log_file_path, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Take user input for the first number
while True:
    num1_input = input("Enter the first number: ")
    if num1_input.strip():  # Check if the input is not empty after removing leading/trailing spaces
        try:
            num1 = float(num1_input)
            break  # Break out of the loop if conversion is successful
        except ValueError:
            print("Invalid input. Please enter a valid number.")
    else:
        print("Input cannot be empty. Please enter a valid number.")

# Take user input for the second number, ensuring it is valid
while True:
    num2_input = input("Enter the second number: ")
    if num2_input.strip():  # Check if the input is not empty after removing leading/trailing spaces
        try:
            num2 = float(num2_input)
            break  # Break out of the loop if conversion is successful
        except ValueError:
            print("Invalid input. Please enter a valid number.")
    else:
        print("Input cannot be empty. Please enter a valid number.")

# Take user input for the mathematical operation
while True:
    operation = input("Enter the operation (+, -, *, /): ")
    if operation.strip() in ['+', '-', '*', '/']:
        break
    else:
        print("Invalid operation. Please enter +, -, *, or /.")

# Perform the operation and log the result
result = perform_operation(num1, num2, operation)

# Path to the Excel file
excel_file_path = os.path.join(excel_directory, 'results.xlsx')

# Initialize or load the Excel workbook
if not os.path.exists(excel_file_path):
    wb = Workbook()
    ws = wb.active
    ws.append(['Operation', 'Result'])
else:
    wb = load_workbook(excel_file_path)
    ws = wb.active

# Append the operation and result to the Excel file
next_row = ws.max_row + 1  # Find the next available row
ws.cell(row=next_row, column=1, value=f"{num1} {operation} {num2}")
ws.cell(row=next_row, column=2, value=result)

# Save the Excel file
wb.save(excel_file_path)

# Display the result
if result is not None:
    print(f"Result of the operation: {result}")

# Path to the SQLite database file
db_file_path = os.path.join(db_directory, 'operations.db')

# Connect to the SQLite database
conn = sqlite3.connect(db_file_path)

# Create the Operations' table if it doesn't exist
create_table(conn)

# Insert the operation and result into the Operations table
if result is not None:
    insert_operation(conn, f"{num1} {operation} {num2}", result)

# Close the database connection
conn.close()
